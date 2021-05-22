from datetime import datetime

import openpyxl

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)
def updateStatus(empid,status,tabs):
    file = r"D:\python practice\Automation\data\Data.xlsx"
    sheetname = "Status"
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]

    import datetime
    date = datetime.date.today()
    testdata1 = (date,str(empid), tabs,status)
    sheet.append(testdata1)
    workbook.save(file)

    # rows=getRowCount(file,sheetname)
    # cols=getColumnCount(file,sheetname)
    #
    # for i in range(1,rows+1):
    #     for j in range(1, cols+1):
    #         print(empid)
    #         print(status)
    #         writeData(file,sheetname,i,j,empid)
    #         writeData(file, sheetname, i, j+1, status)
            # header = readData(file, sheetname, 1, 1)




